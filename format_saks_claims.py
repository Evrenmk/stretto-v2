"""
Saks Filed Claims — JSON to Excel Formatter

Usage: python format_saks_claims.py [input.json] [output.xlsx]

Defaults:
  input:  saks_claims_data.json  (in same directory)
  output: Saks_Filed_Claims.xlsx (in same directory)
"""

import json
import sys
import os
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ──────────────────── Column Definitions ────────────────────

COLUMNS = [
    {"key": "claimNo",                "header": "Claim No.",                  "width": 12,  "type": "number"},
    {"key": "creditorName",           "header": "Creditor Name",             "width": 35,  "type": "text"},
    {"key": "creditorAddress",        "header": "Creditor Address",          "width": 45,  "type": "text"},
    {"key": "debtorName",             "header": "Debtor Name",               "width": 30,  "type": "text"},
    {"key": "dateFiled",              "header": "Date Filed",                "width": 14,  "type": "date"},
    {"key": "claimStatus",            "header": "Claim Status",              "width": 16,  "type": "text"},
    {"key": "scheduleNo",             "header": "Schedule No.",              "width": 14,  "type": "text"},
    {"key": "currentAmountTotal",     "header": "Current Amount (Total)",    "width": 22,  "type": "currency"},
    {"key": "currentGeneralUnsecured","header": "Current - Gen. Unsecured",  "width": 22,  "type": "currency"},
    {"key": "currentPriority",        "header": "Current - Priority",        "width": 18,  "type": "currency"},
    {"key": "currentSecured",         "header": "Current - Secured",         "width": 18,  "type": "currency"},
    {"key": "currentAdminPriority",   "header": "Current - Admin Priority",  "width": 22,  "type": "currency"},
    {"key": "filedAmountTotal",       "header": "Filed Amount (Total)",      "width": 22,  "type": "currency"},
    {"key": "filedGeneralUnsecured",  "header": "Filed - Gen. Unsecured",    "width": 22,  "type": "currency"},
    {"key": "filedPriority",          "header": "Filed - Priority",          "width": 18,  "type": "currency"},
    {"key": "filedSecured",           "header": "Filed - Secured",           "width": 18,  "type": "currency"},
    {"key": "filedAdminPriority",     "header": "Filed - Admin Priority",    "width": 22,  "type": "currency"},
    {"key": "scheduleAmountTotal",    "header": "Schedule Amount (Total)",   "width": 22,  "type": "currency"},
    {"key": "scheduleGeneralUnsecured","header": "Schedule - Gen. Unsecured","width": 22,  "type": "currency"},
    {"key": "schedulePriority",       "header": "Schedule - Priority",       "width": 18,  "type": "currency"},
    {"key": "scheduleSecured",        "header": "Schedule - Secured",        "width": 18,  "type": "currency"},
    {"key": "scheduleAdminPriority",  "header": "Schedule - Admin Priority", "width": 22,  "type": "currency"},
    {"key": "proofOfClaim",           "header": "Proof of Claim (PDF)",      "width": 30,  "type": "link"},
    {"key": "noticeParties",          "header": "Notice Parties",            "width": 30,  "type": "text"},
    {"key": "objectionHistory",       "header": "Objection History",         "width": 30,  "type": "text"},
    {"key": "transferHistory",        "header": "Transfer History",          "width": 30,  "type": "text"},
    {"key": "withdrawalHistory",      "header": "Withdrawal History",        "width": 30,  "type": "text"},
    {"key": "stipulationHistory",     "header": "Stipulation History",       "width": 30,  "type": "text"},
]

# ──────────────────── Styles ────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

ROW_FONT = Font(name="Calibri", size=10)
ROW_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CURRENCY_ALIGNMENT = Alignment(vertical="top", horizontal="right")

EVEN_ROW_FILL = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
ODD_ROW_FILL = PatternFill(fill_type=None)

THIN_BORDER = Border(
    bottom=Side(style="hair", color="D0D0D0")
)

CURRENCY_FORMAT = '#,##0.00'


# ──────────────────── Helpers ────────────────────

def parse_currency(value):
    """Convert '$1,234.56' string to float, or return None."""
    if not value or not isinstance(value, str):
        return None
    cleaned = value.replace('$', '').replace(',', '').strip()
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_claim_number(value):
    """Convert claim number to int for proper sorting."""
    if not value:
        return None
    try:
        return int(str(value).strip())
    except ValueError:
        return value


def format_history(value):
    """Convert JSON history string to readable text."""
    if not value:
        return ""
    try:
        entries = json.loads(value) if isinstance(value, str) else value
        if not entries:
            return ""
        lines = []
        for entry in entries:
            if isinstance(entry, dict):
                parts = [f"{k}: {v}" for k, v in entry.items() if v]
                lines.append("; ".join(parts))
            else:
                lines.append(str(entry))
        return "\n".join(lines)
    except (json.JSONDecodeError, TypeError):
        return str(value)


# ──────────────────── Main ────────────────────

def create_excel(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Filed Claims"

    # Sort by claim number
    data.sort(key=lambda r: (parse_claim_number(r.get("claimNo")) or 99999))

    # ── Header Row ──
    ws.row_dimensions[1].height = 35
    for col_idx, col_def in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

    # ── Data Rows ──
    for row_idx, record in enumerate(data, 2):
        row_fill = EVEN_ROW_FILL if row_idx % 2 == 0 else ODD_ROW_FILL

        for col_idx, col_def in enumerate(COLUMNS, 1):
            raw_value = record.get(col_def["key"], "")
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = ROW_FONT
            cell.border = THIN_BORDER
            cell.fill = row_fill

            col_type = col_def["type"]

            if col_type == "currency":
                numeric = parse_currency(raw_value)
                if numeric is not None:
                    cell.value = numeric
                    cell.number_format = CURRENCY_FORMAT
                    cell.alignment = CURRENCY_ALIGNMENT
                else:
                    cell.value = raw_value or ""
                    cell.alignment = ROW_ALIGNMENT

            elif col_type == "number":
                parsed = parse_claim_number(raw_value)
                cell.value = parsed
                cell.alignment = Alignment(vertical="top", horizontal="center")

            elif col_type == "date":
                cell.value = raw_value or ""
                cell.alignment = Alignment(vertical="top", horizontal="center")

            elif col_type == "link":
                if raw_value:
                    cell.value = raw_value
                    cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                    cell.alignment = ROW_ALIGNMENT
                else:
                    cell.value = ""
                    cell.alignment = ROW_ALIGNMENT

            elif col_type == "text" and col_def["key"].endswith("History"):
                cell.value = format_history(raw_value)
                cell.alignment = ROW_ALIGNMENT

            else:
                cell.value = raw_value or ""
                cell.alignment = ROW_ALIGNMENT

    # ── Freeze & Filter ──
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Save ──
    wb.save(output_path)
    return len(data)


def main():
    script_dir = Path(__file__).parent

    dt = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = script_dir / "output"
    output_dir.mkdir(exist_ok=True)
    default_output = str(output_dir / f"CR_Saks_stretto-parser_{dt}.xlsx")

    input_path = sys.argv[1] if len(sys.argv) > 1 else str(script_dir / "saks_claims_data.json")
    output_path = sys.argv[2] if len(sys.argv) > 2 else default_output

    if not os.path.exists(input_path):
        print(f"Error: Input file not found: {input_path}")
        print(f"Make sure saks_claims_data.json is in {script_dir}")
        sys.exit(1)

    print(f"Reading {input_path}...")
    with open(input_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    print(f"Loaded {len(data)} claims")
    print(f"Formatting Excel with {len(COLUMNS)} columns...")

    count = create_excel(data, output_path)

    print(f"Saved {output_path}")
    print(f"  - {count} claims")
    print(f"  - {len(COLUMNS)} columns")
    print(f"  - Sorted by Claim No.")
    print(f"  - Frozen header with auto-filters")
    print("Done!")


if __name__ == "__main__":
    main()
